#!/usr/bin/env python3

import argparse
from contextlib import nullcontext, contextmanager
import csv
import io
import os
import shutil
import subprocess
import sys
from pathlib import Path
import tempfile
import zipfile
import re
import json
import psutil
import time
from collections import Counter

from openpyxl import load_workbook
from openpyxl.styles.numbers import is_date_format

import x2c_helper as x2c
from x2c_helper import iter_cell_rows, write_csv
from x2c_helper import compose_output_pathname
from x2c_helper import open_output_stream
from x2c_helper import ensure_output_dir
from x2c_helper import touch_stage

# Check if the operating system is Windows
IS_WINDOWS = sys.platform.startswith("win32")
DEFAULT_BACKENDS = ["excel", "libre"] if IS_WINDOWS else ["libre"]

# suffixes of the files to be zipped by default
SUFFIXES_IN_ZIP = [
  ".csv", ".png", ".emf", ".json"
]

SUFFIXES_OUT_ZIP = [
  "-pid.json"
]


def should_include_in_zip(args, workdir_path, target_path):
  if args.debug:
    return True

  workdir_path = Path(workdir_path)
  target_path  = Path(target_path)

  rel = target_path.relative_to(workdir_path)
  if rel.parts[0] == "scratch":
    return False

  if target_path.suffix not in SUFFIXES_IN_ZIP:
      return False

  for s in SUFFIXES_OUT_ZIP:
    if target_path.name.endswith(s):
      return False

  return True


def classify(cell):
  if cell.value is None:
    return "EMPTY"
  if is_date_format(cell.number_format):
    return "DATETIME"
  if cell.number_format == "General":
    return "GENERAL"
  if isinstance(cell.value, (int, float)):
    return "NUMBER"
  return "OTHER"


def iter_format_rows(ws):
  for row in ws.iter_rows():
    yield [classify(cell) for cell in row]


def parse_args():
  parser = argparse.ArgumentParser(
    description="Export XLSX workbook contents into a ZIP file or directory"
  )
  parser.add_argument("xlsx", nargs="*", help="input xlsx file(s)")
  parser.add_argument("--files-from", help="read pathnames of XLSX from file")
  parser.add_argument(
    "--cwd",
    action="store_true",
    help="save auto-generated zip in the current directory instead of input directory",
  )
  parser.add_argument("--dir", help="output directory to emit all ZIP files")
  parser.add_argument("--rawdir", help="output directory to emit all raw CSV (for debugging)")
  parser.add_argument(
    "--xlsx2zip",
    action="store_true",
    help="auto-generate zip file named after the input xlsx",
  )
  parser.add_argument("--zip", help="output zip file path (explicit)")
  parser.add_argument(
    "--emf",
    action="store_true",
    help="export charts as high-resolution EMF files (requires clipboard)",
  )
  parser.add_argument(
    "--cached",
    action="store_true",
    help="write cached.csv using values stored in workbook",
  )
  parser.add_argument(
    "--libre",
    action="store_true",
    help="evaluate formulas with LibreOffice and write value_libre.csv",
  )
  parser.add_argument(
    "--force-value-name",
    action="store_true",
    help="force LibreOffice output to be named 'value.csv' instead of 'value_libre.csv'",
  )
  parser.add_argument(
    "--no-clobber", "-nc",
    action="store_true",
    help="do not clobber existing file",
  )
  parser.add_argument("--excel-timeout", type=int, default=60, help="timeout for Excel worker")
  parser.add_argument("--libre-timeout", type=int, default=10, help="timeout for LibreOffice worker")
  parser.add_argument("--debug", action="store_true", help="debug mode")
  parser.add_argument("--keep-work", action="store_true", help="keep working directory even if zip file is done")
  parser.add_argument("--bom", action="store_true",
    help="add BOM to CSV files",
  )
  parser.add_argument(
    "--backends", type=str, default=",".join(DEFAULT_BACKENDS),
    help="comma separated backend order (default: %(default)s)"
  )
  parser.add_argument("--try-all-backends", action="store_true",
    help="try all backends even after successfully converted"
  )

  args = parser.parse_args()
  args.backends = [
    b
    for b in dict.fromkeys(
      s.strip().lower()
      for s in args.backends.split(",")
    )
  ]
  return args


def create_output_stream(args, dir_path, title, suffix):
  pathname = compose_output_pathname(dir_path, title, suffix)
  return open_output_stream(pathname, bom=args.bom)


def run_worker(args, worker_name, workdir_path, xlsx_path, int_timeout):
  worker_base = Path(worker_name).stem.removeprefix("xlsx2csv-")
  worker_json = workdir_path / (worker_base + "-pid.json")

  cmd = [ sys.executable,
    str(Path(__file__).resolve().parent / worker_name),
    "--dir", str(workdir_path),
    str(xlsx_path),
  ]
  try:
    touch_stage(workdir_path, f"{worker_name}-called-timeout{int_timeout}")
    print(f"call {worker_name}", file=sys.stderr,)
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=int_timeout,)

  except subprocess.TimeoutExpired as e:
    print(f"{worker_name} timed out", file=sys.stderr,)
    touch_stage(workdir_path, f"{worker_name}-expired-timeout")
    print(f"  lookup {worker_json}", file=sys.stderr,)
    if worker_json.exists():
      print(f"  found {worker_json}", file=sys.stderr,)
      touch_stage(workdir_path, f"{worker_base}-pidjson-found")
      worker_info = json.loads( worker_json.read_text(encoding="utf-8") )
      pid = worker_info["pid"]
      ps = psutil.Process(pid)
      if ps.name() == worker_info["name"] and abs(ps.create_time() - worker_info["create_time"]) < 1.0:
        print(f"  kill {pid}", file=sys.stderr,)
        touch_stage(workdir_path, f"{worker_base}-pid{pid}-kill")
        ps.kill()
    return x2c.STATUS_TIMEOUT

  if result.returncode != 0:
    print(f"{worker_name} worker failed: {result.returncode}", file=sys.stderr,)
    print(f"  {result.stderr}", file=sys.stderr,)
    touch_stage(workdir_path, f"{worker_name}-failed")
    return x2c.STATUS_FAILURE

  return x2c.STATUS_SUCCESS


def emit_properties_json(args, wb, workdir_path):
  props = wb.properties
  meta = {}
  for k, v in vars(props).items():
    if v is None:
      meta[k] = None
    elif hasattr(v, "isoformat"):
      meta[k] = v.isoformat()
    else:
      meta[k] = str(v)

  touch_stage(workdir_path, "emit-workbook-properties-json-try")
  Path(workdir_path, "workbook-properties.json").write_text(
    json.dumps(meta, indent=2, ensure_ascii=False, sort_keys=True,),
    encoding="utf-8",
  )
  touch_stage(workdir_path, "emit-workbook-properties-json-done")


@contextmanager
def rec_elapsed(dic, key):
  time_start = time.monotonic()
  try:
    yield
  finally:
    dic[key] = time.monotonic() - time_start


def process_single_xlsx(args, xlsx_path_str):
  xlsx_path = Path(xlsx_path_str).resolve()
  print(f"Processing: {xlsx_path}", file=sys.stderr)

  summary = {}
  summary["xlsx"] = str(xlsx_path.name)
  summary["result"] = False
  summary["backend"] = None

  b_timeout = {
    "excel": args.excel_timeout,
    "libre": args.libre_timeout,
  }
  summary["timeout"] = b_timeout

  b_attempted = []
  summary["backend_attempted"] = b_attempted

  b_result = {}
  summary["backend_result"] = b_result

  timings = {}
  summary["timings"] = timings


  # Determine the target ZIP file path
  zip_path = None
  if args.zip:
    zip_path = Path(args.zip)
  elif args.xlsx2zip:
    if args.cwd:
      zip_path = Path.cwd() / f"{xlsx_path.stem}.zip"
    elif args.dir:
      zip_path = Path(args.dir) / f"{xlsx_path.stem}.zip"
    else:
      zip_path = xlsx_path.with_suffix(".zip")

  if zip_path is not None:
    workdir_path = zip_path.with_suffix(".work")
  elif args.rawdir is not None:
    workdir_path = Path(args.rawdir)
  else:
    return summary

  if args.no_clobber and zip_path.exists():
    print(f"destination {str(zip_path)} is existing, skip conversion", file=sys.stderr)
    return summary

  try:
    with rec_elapsed(timings, "load_workbook"):
      # openpyxl static side (Runs on any OS)
      wb = load_workbook(str(xlsx_path), data_only=False)

    ensure_output_dir(workdir_path, recreate=True)
    emit_properties_json(args, wb, workdir_path)

    with rec_elapsed(timings, "emit_format_formula"):
      for ws in wb.worksheets:
        with create_output_stream(args, workdir_path, ws.title, x2c.SUFFIX_FORMAT) as o:
          write_csv(iter_format_rows(ws), o)

        with create_output_stream(args, workdir_path, ws.title, x2c.SUFFIX_FORMULA) as o:
          write_csv(iter_cell_rows(ws), o)

    if args.cached:
      with rec_elapsed(timings, "emit_cached"):
        wb_cached = load_workbook(str(xlsx_path), data_only=True)
        for ws in wb_cached.worksheets:
          with create_output_stream(args, workdir_path, ws.title, x2c.SUFFIX_CACHED) as o:
            write_csv(iter_cell_rows(ws), o)


    backend_result = {}
    for b in args.backends:
      b_attempted.append(b)
      evb = "emit_value_" + b
      with rec_elapsed(timings, evb):
        b_result[b] = run_worker( args, f"xlsx2csv-{b}.py", workdir_path, xlsx_path, b_timeout[b])

      if b_result[b]:
        summary["result"] = True
        if summary["backend"] is None:
          summary["backend"] = b
        if args.try_all_backends:
          continue
        break

    print("write summary.json", file=sys.stderr)
    Path(workdir_path, "summary.json").write_text(
      json.dumps(summary, indent=2, ensure_ascii=False, sort_keys=False,),
      encoding="utf-8",
    )

    if not summary["result"]:
      return summary

    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED,) as zf:
      # print(f"  {zip_path} packing started", file=sys.stderr)
      touch_stage(workdir_path, "zip-packing-started")
      for path in Path(workdir_path).rglob("*"):
        rel_path = path.relative_to(workdir_path)
        # print(f"  test {path}", file=sys.stderr)
        if not should_include_in_zip(args, workdir_path, path):
          # print(f"  {zip_path.name} excludes {rel_path}", file=sys.stderr)
          continue
        print(f"  {zip_path.name} includes {rel_path}", file=sys.stderr)
        zf.write(path, rel_path,)
      touch_stage(workdir_path, "zip-packing-done")
      print(f"  {zip_path} completed", file=sys.stderr)

    if not args.keep_work:
      print(f"  remove {workdir_path}", file=sys.stderr)
      shutil.rmtree(workdir_path)

  except Exception as e:
    print("something goes wrong", file=sys.stderr)

  finally:
    print("finished\n", file=sys.stderr)

  return summary

def summarize(summaries):
  backend_counter = Counter()
  timing_collection = {}
  for xlsx,summary in summaries.items():
    timings = summary["timings"]

    for k,v in timings.items():
      if k.startswith("emit_value_"):
        continue
      elif k not in timing_collection:
        timing_collection[k] = []
      timing_collection[k].append(v)

    for b,result in summary["backend_result"].items():
      if not result:
        continue

      backend_counter[b] += 1

      evb = "emit_value_" + b
      if evb not in timing_collection:
        timing_collection[evb] = []
      timing_collection[evb].append(timings[evb])

  print(f"summary of {len(summaries)} files", file=sys.stderr)
  print("backend", file=sys.stderr)
  for k,v in backend_counter.items():
    print(f"  {k}={v}")
  print("timings", file=sys.stderr)
  for k,v in timing_collection.items():
    print(f"  timing {k}: avg={sum(v)/len(v):.2f} max={max(v):.2f}", file=sys.stderr)


def main():
  args = parse_args()
  summaries = {}

  if args.files_from:
    ctx = (
      nullcontext(sys.stdin)
      if args.files_from == "-"
      else open(args.files_from, "r", encoding="utf-8")
    )
    with ctx as fh:
      for xlsx_target in fh:
        xlsx_target = xlsx_target.rstrip("\r\n")
        print(f"open file {xlsx_target} specified by file list", file=sys.stderr)
        summary = process_single_xlsx(args, xlsx_target)
        summaries[xlsx_target] = summary
        summarize(summaries)
  else:
    print(args.xlsx)
    # Iterate over all provided XLSX targets
    for xlsx_target in args.xlsx:
      print(f"open file {xlsx_target} specified by argument", file=sys.stderr)
      summary = process_single_xlsx(args, xlsx_target)
      summaries[xlsx_target] = summary
      summarize(summaries)



if __name__ == "__main__":
  main()
